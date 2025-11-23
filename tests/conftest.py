import os
import re
import csv
import pytest
from dotenv import load_dotenv
from openpyxl import load_workbook
import requests
from google_sheets import append_rows_to_sheet



load_dotenv()


def _load_forms_urls():
    raw = os.getenv("PPROFIT_URLS", "").strip()
    if not raw:
        pytest.skip("No URLs provided via PPROFIT_URLS for test_forms")

    urls: list[str] = []

    if os.path.exists(raw):
        _, ext = os.path.splitext(raw)
        ext = ext.lower()
        try:
            if ext in (".xlsx", ".xls"):
                wb = load_workbook(raw, read_only=True, data_only=True)
                try:
                    ws = wb.active
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if isinstance(cell, str):
                                val = cell.strip()
                                if val.startswith("http://") or val.startswith("https://"):
                                    urls.append(val)
                finally:
                    wb.close()
            elif ext == ".csv":
                with open(raw, newline="", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    for row in reader:
                        for cell in row:
                            val = (cell or "").strip()
                            if val.startswith("http://") or val.startswith("https://"):
                                urls.append(val)
            else:
                with open(raw, encoding="utf-8") as f:
                    content = f.read()
                for token in re.split(r"[\s,]+", content):
                    val = token.strip()
                    if val and (val.startswith("http://") or val.startswith("https://")):
                        urls.append(val)
        except Exception as e:
            raise pytest.UsageError(f"Failed to read URLs from file {raw}: {e}")
    else:
        urls = [u.strip() for u in raw.split(",") if u.strip()]

    if not urls:
        pytest.skip("No valid http/https URLs found in PPROFIT_URLS source")
    return urls


@pytest.fixture(scope="session", autouse=True)
def _telegram_notify_on_finish(request):
    """
    По окончании всех тестов отправляет оповещение в Telegram с ссылкой на отчет Google Sheets,
    если заданы BOT_TOKEN, CHAT_ID и SPREADSHEET_ID/GOOGLE_SHEETS_SPREADSHEET_ID
    """
    yield
    # Батч-запись собранных результатов
    try:
        batch_rows = request.session._collected_rows if hasattr(request.session, "_collected_rows") else []
        if batch_rows:
            spreadsheet_id = (os.getenv("GOOGLE_SHEETS_SPREADSHEET_ID") or os.getenv("SPREADSHEET_ID") or "").strip()
            worksheet_name = (os.getenv("GOOGLE_SHEETS_WORKSHEET") or "Итоги3").strip() or "Итоги3"
            if spreadsheet_id:
                append_rows_to_sheet(spreadsheet_id=spreadsheet_id, worksheet_name=worksheet_name, rows=batch_rows)
    except Exception:
        pass
    # Only controller process (no workerinput) should send Telegram messages to avoid duplicates under xdist
    if hasattr(request.config, "workerinput"):
        return

    bot_token = (os.getenv("BOT_TOKEN") or "").strip()
    chat_id = (os.getenv("CHAT_ID") or "").strip()
    spreadsheet_id = (os.getenv("GOOGLE_SHEETS_SPREADSHEET_ID") or os.getenv("SPREADSHEET_ID") or "").strip()
    report_url_override = (os.getenv("TELEGRAM_REPORT_URL") or "").strip()
    if not (bot_token and chat_id and (spreadsheet_id or report_url_override)):
        return

    # Сформируем ссылку на Google Sheets: приоритет у TELEGRAM_REPORT_URL, иначе строим из ID
    if report_url_override:
        link = report_url_override
    else:
        link = spreadsheet_id
        if "/d/" not in link:
            link = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"

    text = f"Парсер по проверке улиц закончил работу. Сссылка на отчет - {link}"
    try:
        requests.post(
            f"https://api.telegram.org/bot{bot_token}/sendMessage",
            json={"chat_id": chat_id, "text": text},
            timeout=10,
        )
    except Exception:
        pass


@pytest.fixture(params=_load_forms_urls())
def example_url(request):
    return request.param


@pytest.fixture(scope="function")
def collect_row(request):
    """Собирает строки для батч-записи в Google Sheets в конце сессии"""
    if not hasattr(request.session, "_collected_rows"):
        request.session._collected_rows = []
    def _add(row: list[str]):
        request.session._collected_rows.append(row)
    return _add